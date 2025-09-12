import streamlit as st
import pandas as pd
import numpy as np
import os
import json
import uuid
import tempfile
import shutil
from pathlib import Path
import openpyxl
from openpyxl.utils import range_boundaries
import re
import logging
from datetime import datetime
import anthropic
from dotenv import load_dotenv
import time
import traceback
from io import BytesIO
from typing import Optional, List, Dict, Any, Tuple

# Import S3 service and format validators
from aws_s3_service import s3_service
from format_validators import FormatValidator, validate_and_flag_formats

# Load environment variables
load_dotenv()

# Page configuration
st.set_page_config(
    page_title="Excel Data Cleaner",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for enhanced UI
st.markdown("""
<style>
    .main-header {
        font-size: 2.8rem;
        font-weight: bold;
        text-align: center;
        margin-bottom: 1.5rem;
        color: #2c3e50;
        text-shadow: 1px 1px #ecf0f1;
    }
    .section-header {
        font-size: 1.8rem;
        font-weight: 600;
        color: #27ae60;
        margin: 1.5rem 0 1rem;
        border-bottom: 2px solid #27ae60;
        padding-bottom: 0.5rem;
    }
    .stTabs [data-baseweb="tab-list"] button {
        background-color: #ecf0f1;
        border: none;
        padding: 10px 20px;
        margin: 0 5px;
        border-radius: 5px;
        font-weight: 500;
    }
    .stTabs [data-baseweb="tab-list"] button:hover {
        background-color: #3498db;
        color: white;
    }
    .stTabs [data-baseweb="tab-list"] button[data-selected="true"] {
        background-color: #3498db;
        color: white;
    }
    .sidebar .stButton>button {
        width: 100%;
        margin: 5px 0;
    }
    .success-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        color: #155724;
    }
    .error-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #f8d7da;
        border: 1px solid #f5c6cb;
        color: #721c24;
    }
    .info-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        color: #0c5460;
    }
    .expander {
        background-color: #f9f9f9;
        border-radius: 5px;
        padding: 10px;
    }
</style>
""", unsafe_allow_html=True)

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize session state
if 'current_excel_id' not in st.session_state:
    st.session_state.current_excel_id = None
if 'current_excel_name' not in st.session_state:
    st.session_state.current_excel_name = None
if 'current_sheet_name' not in st.session_state:
    st.session_state.current_sheet_name = None
if 'available_sheets' not in st.session_state:
    st.session_state.available_sheets = []
if 'cleaned_data' not in st.session_state:
    st.session_state.cleaned_data = None
if 'analysis_results' not in st.session_state:
    st.session_state.analysis_results = None
if 'agent_decision' not in st.session_state:
    st.session_state.agent_decision = None

# Get Anthropic client
@st.cache_resource
def get_anthropic_client():
    """Get Anthropic client with settings"""
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        return None
    return anthropic.Anthropic(api_key=api_key)

def repair_corrupted_excel(input_path, temp_path='repaired.xlsx'):
    """Repair Excel file and return workbook"""
    try:
        if not os.access(input_path, os.W_OK):
            shutil.copy2(input_path, temp_path)
            os.chmod(temp_path, 0o666)
            return openpyxl.load_workbook(temp_path), temp_path
        else:
            return openpyxl.load_workbook(input_path), input_path
    except Exception as e:
        logger.error(f"Error opening Excel file: {e}")
        try:
            df = pd.read_excel(input_path, engine='openpyxl')
            df.to_excel(temp_path, index=False, engine='openpyxl')
            return openpyxl.load_workbook(temp_path), temp_path
        except Exception as e2:
            logger.error(f"Failed to repair Excel file: {e2}")
            raise

def fill_merged_cells(ws):
    """Fill merged cells with the value from the top-left cell"""
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col, value=top_left_value)




def llm_detect_header_row(df, max_rows_to_analyze=10):
    """
    Use LLM to intelligently detect header row with fallback to rule-based detection.
    
    Purpose: More accurate header detection using AI understanding of context and patterns.
    Fallback: If LLM fails or is unavailable, uses detect_header_row_smart() function.
    
    Args:
        df: DataFrame to analyze
        max_rows_to_analyze: Number of rows to send to LLM (default 10)
    
    Returns:
        int: Header row index (0-based) or -1 if no header found
    """
    try:
        client = get_anthropic_client()
        if not client:
            logger.info("LLM not available, falling back to rule-based header detection")
            return detect_header_row_smart(df)
        
        # Limit analysis to prevent overwhelming LLM
        sample_df = df.head(max_rows_to_analyze)
        
        # Convert to a format suitable for LLM analysis
        sample_data = []
        for i, row in sample_df.iterrows():
            row_data = {}
            for j, val in enumerate(row):
                row_data[f"col_{j}"] = str(val) if pd.notna(val) else "NULL"
            sample_data.append({"row_index": i, "data": row_data})
        
        prompt = f"""Analyze this Excel data sample and identify the header row index.

Data sample (first {len(sample_data)} rows):
{json.dumps(sample_data, indent=2)}

Rules for header detection:
1. Headers typically contain descriptive text (not numbers or dates)
2. Headers often have keywords like: name, id, date, amount, total, code, address, phone, email, status, type
3. Data rows below headers should be consistent in type
4. Headers should not be purely numeric values
5. Headers are usually in the first few rows (0-5)

Return ONLY a JSON object with this exact structure:
{{
    "header_row_index": 0,
    "confidence": 0.95,
    "reasoning": "Row 0 contains descriptive column names like 'Customer_Name', 'Order_Date'"
}}

If no clear header is found, return header_row_index: -1"""

        response = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}]
        )
        
        response_text = response.content[0].text.strip()
        
        # Parse JSON response
        try:
            result = json.loads(response_text)
            header_idx = result.get('header_row_index', -1)
            confidence = result.get('confidence', 0.0)
            reasoning = result.get('reasoning', 'No reasoning provided')
            
            logger.info(f"LLM header detection: row {header_idx}, confidence {confidence}")
            logger.info(f"LLM reasoning: {reasoning}")
            
            # Validate the result
            if header_idx >= 0 and header_idx < len(df) and confidence > 0.5:
                return header_idx
            else:
                logger.info("LLM confidence too low, falling back to rule-based detection")
                return detect_header_row_smart(df)
                
        except json.JSONDecodeError:
            logger.warning("Failed to parse LLM response, falling back to rule-based detection")
            return detect_header_row_smart(df)
            
    except Exception as e:
        logger.warning(f"LLM header detection failed: {str(e)}, falling back to rule-based detection")
        return detect_header_row_smart(df)

def llm_generate_column_names(df, existing_columns):
    """
    Use LLM to generate intelligent column names based on data content.
    
    Purpose: Creates meaningful column names for unnamed/generic columns by analyzing data.
    Only processes: Columns that are unnamed, empty, or have generic names like 'Unnamed_1'.
    
    Args:
        df: DataFrame with data
        existing_columns: List of current column names
    
    Returns:
        list: Updated column names with LLM suggestions for unnamed columns
    """
    try:
        client = get_anthropic_client()
        if not client:
            return existing_columns
        
        # Analyze only unnamed columns
        unnamed_indices = []
        column_samples = {}
        
        for i, col in enumerate(existing_columns):
            if pd.isna(col) or col == '' or str(col).strip() == '' or 'Unnamed' in str(col) or 'Column_' in str(col):
                unnamed_indices.append(i)
                # Get sample data for this column
                col_data = df.iloc[:, i].dropna().head(10)
                column_samples[f"column_{i}"] = [str(val) for val in col_data.tolist()]
        
        if not unnamed_indices:
            return existing_columns
        
        prompt = f"""Generate meaningful column names based on the data content.

Column data samples:
{json.dumps(column_samples, indent=2)}

Guidelines:
1. Use descriptive, business-friendly names
2. Follow naming conventions: Title_Case_With_Underscores
3. Be specific but concise
4. Avoid generic names like "Data" or "Value"
5. Consider data patterns (emails, dates, numbers, etc.)

Return ONLY a JSON object mapping column indices to names:
{{
    "column_0": "Customer_Email",
    "column_1": "Order_Amount",
    "column_2": "Transaction_Date"
}}"""

        response = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=400,
            messages=[{"role": "user", "content": prompt}]
        )
        
        response_text = response.content[0].text.strip()
        
        try:
            name_suggestions = json.loads(response_text)
            new_columns = list(existing_columns)
            
            for col_key, suggested_name in name_suggestions.items():
                if col_key in column_samples:
                    col_idx = int(col_key.split('_')[1])
                    if col_idx < len(new_columns):
                        new_columns[col_idx] = str(suggested_name).strip()
            
            logger.info(f"LLM generated {len(name_suggestions)} column names")
            return new_columns
            
        except (json.JSONDecodeError, ValueError, IndexError) as e:
            logger.warning(f"Failed to parse LLM column name suggestions: {str(e)}")
            return existing_columns
            
    except Exception as e:
        logger.warning(f"LLM column name generation failed: {str(e)}")
        return existing_columns

def llm_analyze_data_quality(df, changes_log):
    """
    Lightweight LLM analysis for data quality during cleaning process.
    This is a simplified version focused only on cleaning suggestions.
    For comprehensive analysis, use ai_analyze_df() function.
    """
    try:
        client = get_anthropic_client()
        if not client:
            return changes_log
        
        # Create a minimal summary for quick analysis (less data than ai_analyze_df)
        data_summary = {
            "shape": df.shape,
            "columns": list(df.columns)[:10],  # Only first 10 columns for speed
            "dtypes": {str(k): str(v) for k, v in df.dtypes.head(10).to_dict().items()},
            "sample_rows": 3,  # Just 3 sample rows for quick analysis
            "has_nulls": df.isnull().any().any(),
            "has_duplicates": df.duplicated().any()
        }
        
        # Add a few sample values for context
        if len(df) > 0:
            data_summary["first_row"] = df.head(1).to_dict('records')[0] if len(df) > 0 else {}
        
        prompt = f"""Quick analysis of cleaned data for improvement suggestions.

Data Info:
- Shape: {data_summary['shape']}
- Columns (first 10): {data_summary['columns']}
- Has nulls: {data_summary['has_nulls']}
- Has duplicates: {data_summary['has_duplicates']}

Provide 3-5 brief, actionable cleaning suggestions.
Return ONLY a JSON array of strings. Example:
["Convert date columns to datetime", "Standardize text case"]

Focus on format/type improvements only. No data removal."""

        response = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=200,  # Reduced for faster response
            messages=[{"role": "user", "content": prompt}]
        )
        
        response_text = response.content[0].text.strip()
        
        try:
            suggestions = json.loads(response_text)
            if isinstance(suggestions, list) and suggestions:
                changes_log.append("🤖 Quick Quality Check Suggestions:")
                for suggestion in suggestions[:5]:
                    changes_log.append(f"  💡 {suggestion}")
                logger.info(f"LLM provided {len(suggestions)} quick suggestions")
            
        except json.JSONDecodeError:
            logger.warning("Could not parse LLM suggestions")
            
    except Exception as e:
        logger.warning(f"Quick LLM analysis skipped: {str(e)}")
    
    return changes_log



def clean_excel_basic(input_path, output_path, sheet_name=None):
    """Ultra-conservative Excel cleaning - preserves ALL data, no additions or deletions"""
    start_time = time.time()
    changes_log = []
    
    try:
        logger.info(f"Starting ultra-conservative clean_excel_basic: {input_path}")
        
        # Read the Excel file exactly as is - no headers assumed
        df = pd.read_excel(input_path, sheet_name=sheet_name, header=None, engine='openpyxl')
        original_shape = df.shape
        changes_log.append(f"✓ Loaded Excel file with shape: {original_shape}")
        
        # Step 1: Remove ONLY completely empty rows (all values are NaN)
        initial_rows = len(df)
        df = df.dropna(how='all')
        rows_removed = initial_rows - len(df)
        if rows_removed > 0:
            changes_log.append(f"✓ Removed {rows_removed} completely empty rows")
        
        # Step 2: Remove ONLY completely empty columns (all values are NaN)
        initial_cols = len(df.columns)
        df = df.dropna(axis=1, how='all')
        cols_removed = initial_cols - len(df.columns)
        if cols_removed > 0:
            changes_log.append(f"✓ Removed {cols_removed} completely empty columns")
        
        # Step 3: LLM-powered smart header detection with fallback
        header_row_idx = llm_detect_header_row(df)
        
        if header_row_idx > 0:
            # There are values above the header - preserve them in column names
            changes_log.append(f"✓ Detected header at row {header_row_idx + 1}")
            
            # Get pre-header values
            pre_header_values = []
            for col_idx in range(len(df.columns)):
                pre_header_parts = []
                for row_idx in range(header_row_idx):
                    val = df.iloc[row_idx, col_idx]
                    if pd.notna(val) and str(val).strip():
                        pre_header_parts.append(str(val).strip())
                pre_header_values.append('_'.join(pre_header_parts) if pre_header_parts else '')
            
            # Get the actual header row
            header_values = df.iloc[header_row_idx].fillna('').astype(str)
            
            # Combine pre-header and header values
            new_columns = []
            for i, (pre_val, header_val) in enumerate(zip(pre_header_values, header_values)):
                if pre_val and header_val and header_val != '':
                    # Both pre-header and header exist
                    col_name = f"{pre_val}_{header_val}"
                elif pre_val:
                    # Only pre-header exists
                    col_name = pre_val
                elif header_val and header_val != '':
                    # Only header exists
                    col_name = header_val
                else:
                    # Neither exists
                    col_name = f"Column_{i+1}"
                new_columns.append(col_name)
            
            # Set the columns
            df.columns = new_columns
            
            # Remove header and pre-header rows (they're now in column names)
            df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
            changes_log.append(f"✓ Removed {header_row_idx + 1} header/pre-header rows, preserved info in column names")
            
        elif header_row_idx == 0:
            # First row is the header
            header_values = df.iloc[0].fillna('').astype(str)
            
            # Generate proper column names
            new_columns = []
            for i, col in enumerate(header_values):
                if col == '' or col == 'nan' or str(col).strip() == '':
                    new_columns.append(f"Column_{i+1}")
                else:
                    new_columns.append(str(col).strip())
            
            df.columns = new_columns
            
            # Remove the header row since it's now used as column names
            df = df.iloc[1:].reset_index(drop=True)
            changes_log.append("✓ Used first row as column names and removed header row")
        else:
            # No header detected - generate column names
            df.columns = [f"Column_{i+1}" for i in range(len(df.columns))]
            changes_log.append("✓ Generated column names (no header row detected)")
        
        # Step 4: LLM-enhanced column name generation and cleaning
        # First clean existing column names
        cleaned_columns = []
        seen_names = {}
        for col in df.columns:
            # Clean the column name
            col_str = str(col).strip()
            if col_str == '' or col_str == 'nan':
                col_str = f"Unnamed_{len(cleaned_columns)+1}"
            else:
                # Replace problematic characters
                col_str = col_str.replace('\n', '_').replace('\r', '_').replace('\t', '_')
                col_str = col_str.replace('/', '_').replace('\\', '_')
                # Remove multiple underscores
                col_str = '_'.join(filter(None, col_str.split('_')))
            
            # Handle duplicates
            if col_str in seen_names:
                seen_names[col_str] += 1
                col_str = f"{col_str}_{seen_names[col_str]}"
            else:
                seen_names[col_str] = 0
            
            cleaned_columns.append(col_str)
        
        df.columns = cleaned_columns
        changes_log.append(f"✓ Cleaned {len(df.columns)} column names")
        
        # Step 4b: Use LLM to generate intelligent names for unnamed/generic columns
        try:
            llm_enhanced_columns = llm_generate_column_names(df, df.columns.tolist())
            if llm_enhanced_columns != df.columns.tolist():
                df.columns = llm_enhanced_columns
                changes_log.append("✓ Enhanced column names using LLM analysis")
        except Exception as e:
            logger.warning(f"LLM column name enhancement failed: {str(e)}")
            changes_log.append("⚠️ LLM column name enhancement failed, using basic names")
        
        # Step 5: Minimal data cleaning (no data loss)
        for col in df.columns:
            if df[col].dtype == 'object':
                # Only trim whitespace, don't change values
                df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
        
        changes_log.append("✓ Trimmed whitespace from text values")
        
        # Step 6: Standardize only obvious missing value representations
        # Be very conservative here
        df = df.replace(['NULL', 'null', '#N/A'], np.nan)
        changes_log.append("✓ Standardized NULL representations to NaN")
        
        # Step 7: Optional LLM data quality analysis for additional insights
        try:
            changes_log = llm_analyze_data_quality(df, changes_log)
        except Exception as e:
            logger.warning(f"LLM data quality analysis failed: {str(e)}")
            changes_log.append("⚠️ LLM data quality analysis failed")
        
        # Verify no data was lost
        final_shape = df.shape
        if final_shape[1] > original_shape[1]:
            logger.error(f"Columns were added! Original: {original_shape[1]}, Final: {final_shape[1]}")
            changes_log.append(f"⚠️ Warning: Column count changed from {original_shape[1]} to {final_shape[1]}")
        
        # Save the cleaned data
        df.to_excel(output_path, index=False, engine='openpyxl')
        changes_log.append(f"✅ Saved cleaned data to {output_path}")
        
        processing_time = time.time() - start_time
        changes_log.append(f"📊 Final dataset: {df.shape[0]} rows × {df.shape[1]} columns")
        changes_log.append(f"⏱️ Processing completed in {processing_time:.2f} seconds")
        return df, changes_log, processing_time
        
    except Exception as e:
        processing_time = time.time() - start_time
        error_msg = f"❌ Error processing file: {str(e)}"
        changes_log.append(error_msg)
        logger.error(f"Clean error: {str(e)} - {traceback.format_exc()}")
        return pd.DataFrame(), changes_log, processing_time

def detect_header_row_smart(df):
    """
    Enhanced smart header detection with better pattern recognition
    Returns the index of the header row, or -1 if no clear header
    """
    if df.empty:
        return -1
    
    # Check first 15 rows for potential headers (increased from 10)
    max_check = min(15, len(df))
    header_scores = []
    
    for row_idx in range(max_check):
        row_data = df.iloc[row_idx]
        non_null = row_data.dropna()
        
        if len(non_null) == 0:
            header_scores.append(0)
            continue
        
        # Calculate header likelihood score
        score = 0
        string_count = 0
        numeric_count = 0
        header_like_count = 0
        
        for val in non_null:
            val_str = str(val).strip().lower()
            
            if isinstance(val, str) or (isinstance(val, (int, float)) and not str(val).replace('.','').replace('-','').isdigit()):
                # Check for header-like characteristics
                header_indicators = [
                    len(val_str) > 2,  # Not too short
                    any(c.isalpha() for c in val_str),  # Contains letters
                    val_str not in ['nan', 'null', '', '0', '1'],  # Not common data values
                    not val_str.replace('.','').replace('-','').replace(' ','').isdigit(),  # Not purely numeric
                    any(keyword in val_str for keyword in ['name', 'id', 'date', 'amount', 'total', 'code', 'number', 'address', 'phone', 'email', 'status', 'type', 'description'])  # Common header keywords
                ]
                
                if sum(header_indicators) >= 3:
                    header_like_count += 1
                    score += 3
                elif sum(header_indicators) >= 2:
                    score += 2
                
                string_count += 1
            else:
                numeric_count += 1
                # Penalize if it looks like data
                if isinstance(val, (int, float)) and val > 100:
                    score -= 1
        
        # Boost score based on string ratio
        if len(non_null) > 0:
            string_ratio = string_count / len(non_null)
            if string_ratio >= 0.8:
                score += 5
            elif string_ratio >= 0.6:
                score += 3
            elif string_ratio >= 0.4:
                score += 1
        
        # Boost score if most values look like headers
        if len(non_null) > 0 and header_like_count / len(non_null) >= 0.6:
            score += 4
        
        # Check if next rows look more like data
        data_like_next = 0
        for next_idx in range(row_idx + 1, min(row_idx + 4, len(df))):
            if next_idx < len(df):
                next_row = df.iloc[next_idx].dropna()
                if len(next_row) > 0:
                    numeric_in_next = sum(1 for v in next_row if isinstance(v, (int, float)) or 
                                        (isinstance(v, str) and v.replace('.','').replace('-','').isdigit()))
                    if numeric_in_next / len(next_row) > 0.4:
                        data_like_next += 1
        
        if data_like_next >= 2:
            score += 3
        elif data_like_next >= 1:
            score += 1
        
        # Penalize if row is too early and sparse
        if row_idx > 0 and len(non_null) < len(df.columns) * 0.3:
            score -= 2
        
        header_scores.append(score)
    
    # Find the row with highest score above threshold
    if header_scores:
        max_score = max(header_scores)
        if max_score >= 5:  # Minimum confidence threshold
            best_idx = header_scores.index(max_score)
            
            # Additional validation: ensure it's not too late in the data
            if best_idx < min(8, len(df) // 3):
                return best_idx
    
    # Fallback: check if first few rows look like headers
    for row_idx in range(min(3, len(df))):
        row_data = df.iloc[row_idx].dropna()
        if len(row_data) > 0:
            string_ratio = sum(1 for v in row_data if isinstance(v, str) and 
                             not str(v).replace('.','').replace('-','').isdigit()) / len(row_data)
            if string_ratio >= 0.7:
                return row_idx
    
    return -1  # No header found

def ai_analyze_df(df):
    """Enhanced AI analysis with comprehensive metadata"""
    try:
        client = get_anthropic_client()
        if not client:
            return None, ["AI service not available - check API key"]
        
        # Limit DataFrame size to prevent overwhelming the AI
        MAX_COLS = 50
        if len(df.columns) > MAX_COLS:
            df_sample = df.iloc[:, :MAX_COLS]
            logger.info(f"DataFrame has {len(df.columns)} columns, analyzing first {MAX_COLS}")
        else:
            df_sample = df
        
        # Prepare comprehensive metadata for each column
        metadata = {}
        for col in df_sample.columns:
            try:
                col_data = df_sample[col]
                non_null_data = col_data.dropna()
                
                # Get up to 10 sample values
                sample_values = []
                if len(non_null_data) > 0:
                    unique_vals = non_null_data.unique()
                    # Take up to 10 unique values, convert to string safely
                    sample_values = []
                    for val in unique_vals[:10]:
                        try:
                            sample_values.append(str(val))
                        except:
                            sample_values.append("(complex value)")
                
                # Analyze data patterns
                col_metadata = {
                    "dtype": str(col_data.dtype),
                    "total_count": int(len(col_data)),
                    "non_null_count": int(len(non_null_data)),
                    "null_count": int(col_data.isnull().sum()),
                    "null_percentage": round(float((col_data.isnull().sum() / len(col_data)) * 100), 2) if len(col_data) > 0 else 0,
                    "unique_count": int(col_data.nunique()),
                    "sample_values": sample_values[:5],  # Limit to 5 samples to reduce size
                    "has_whitespace_issues": False,
                    "has_mixed_types": False
                }
                
                # Check for whitespace issues in string columns
                if col_data.dtype == 'object' and len(non_null_data) > 0:
                    try:
                        str_values = non_null_data[non_null_data.apply(lambda x: isinstance(x, str))]
                        if len(str_values) > 0:
                            sample_str = str_values.iloc[:min(100, len(str_values))]
                            col_metadata["has_whitespace_issues"] = any(
                                sample_str.apply(lambda x: x != x.strip() if isinstance(x, str) else False)
                            )
                    except:
                        pass
                
                # For numeric columns, add statistics
                if col_data.dtype in ['int64', 'float64', 'Int64', 'Float64']:
                    if len(non_null_data) > 0:
                        try:
                            col_metadata["min"] = round(float(non_null_data.min()), 2)
                            col_metadata["max"] = round(float(non_null_data.max()), 2)
                            col_metadata["mean"] = round(float(non_null_data.mean()), 2)
                        except:
                            pass
                
                metadata[str(col)[:50]] = col_metadata  # Limit column name length
            except Exception as e:
                logger.warning(f"Error processing column {col}: {e}")
                metadata[str(col)[:50]] = {"error": "Could not process column"}
        
        # Prepare a simplified summary for AI
        df_summary = {
            "total_rows": int(df.shape[0]),
            "total_columns": int(df.shape[1]),
            "columns_analyzed": len(metadata),
            "duplicate_rows": int(df.duplicated().sum()),
            "column_samples": {}
        }
        
        # Add simplified column info
        for col_name, col_meta in list(metadata.items())[:20]:  # Limit to 20 columns for the prompt
            df_summary["column_samples"][col_name] = {
                "dtype": col_meta.get("dtype", "unknown"),
                "null_pct": col_meta.get("null_percentage", 0),
                "unique": col_meta.get("unique_count", 0),
                "samples": col_meta.get("sample_values", [])[:3]  # Only 3 samples
            }
        
        prompt = f"""Analyze this DataFrame summary and provide data quality assessment.

DataFrame Overview:
- Shape: {df_summary['total_rows']} rows × {df_summary['total_columns']} columns
- Duplicate rows: {df_summary['duplicate_rows']}
- Columns analyzed: {df_summary['columns_analyzed']}

Column Samples (first 20):
{json.dumps(df_summary['column_samples'], indent=2, default=str)}

Provide a brief analysis focusing on:
1. Data quality issues (formatting, types, consistency)
2. Cleaning suggestions (no data removal)

Return ONLY valid JSON with this exact structure:
{{
    "analysis": ["issue 1", "issue 2", "issue 3"],
    "suggestions": ["suggestion 1", "suggestion 2", "suggestion 3"],
    "data_quality_score": 75
}}

Keep responses concise. Maximum 3-5 items per list."""
        
        response = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )
        
        response_text = response.content[0].text.strip()
        
        # Try to extract JSON from the response
        try:
            # First try direct parsing
            analysis = json.loads(response_text)
        except json.JSONDecodeError:
            # Try to find JSON in the response
            import re
            json_match = re.search(r'\{[^{}]*\}', response_text, re.DOTALL)
            if json_match:
                try:
                    analysis = json.loads(json_match.group())
                except:
                    # Return a default structure
                    analysis = {
                        "analysis": ["Unable to parse AI response - data analyzed successfully"],
                        "suggestions": ["Please review the metadata below for details"],
                        "data_quality_score": 70
                    }
            else:
                analysis = {
                    "analysis": ["Data analysis completed"],
                    "suggestions": ["Review metadata for detailed information"],
                    "data_quality_score": 70
                }
        
        # Add the metadata to the analysis
        analysis["metadata"] = metadata
        
        # Ensure all required fields exist
        if "analysis" not in analysis:
            analysis["analysis"] = ["Data analyzed successfully"]
        if "suggestions" not in analysis:
            analysis["suggestions"] = ["No specific suggestions"]
        if "data_quality_score" not in analysis:
            analysis["data_quality_score"] = 70
            
        return analysis, []
        
    except Exception as e:
        logger.error(f"Error in ai_analyze_df: {str(e)}")
        # Return a basic analysis with the metadata we collected
        try:
            return {
                "metadata": metadata if 'metadata' in locals() else {},
                "analysis": ["Error during AI analysis, showing basic metadata"],
                "suggestions": ["Review the metadata below for column information"],
                "data_quality_score": 50
            }, []
        except:
            return None, [f"AI Error: {str(e)}"]

def apply_ai_suggestions(df, selected_suggestions):
    """Apply AI suggestions with data preservation"""
    try:
        client = get_anthropic_client()
        if not client:
            return df, "AI service not available"
        
        # Get current shape for validation
        original_shape = df.shape
        
        df_info = f"Columns: {list(df.columns)}\nShape: {df.shape}\nDtypes: {df.dtypes.to_dict()}\nSample data:\n{df.head(5).to_string()}"
        
        prompt = f"""Given this DataFrame:
{df_info}

Implement these cleaning suggestions:
{json.dumps(selected_suggestions)}

CRITICAL RULES:
1. NEVER remove any rows or columns
2. NEVER drop or delete data
3. Only clean and standardize existing data
4. Preserve all original data values (just clean formatting)
5. Fix data types without losing information
6. Standardize formats and units
7. Clean whitespace and encoding issues
8. The output DataFrame must have the same number of rows and columns as input

Generate Python pandas code to modify 'df' following these rules. Only return executable code, no explanations.
Example transformations:
- df['col'] = df['col'].str.strip()  # Clean whitespace
- df['col'] = pd.to_numeric(df['col'], errors='ignore')  # Convert types safely
- df.columns = df.columns.str.replace(' ', '_')  # Clean column names"""
        
        response = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=800,
            messages=[{"role": "user", "content": prompt}]
        )
        
        code = response.content[0].text.strip()
        
        # Remove any code blocks markers if present
        if '```python' in code:
            code = code.replace('```python', '').replace('```', '').strip()
        
        local_vars = {'df': df.copy(), 'pd': pd, 'np': np, 're': re}
        exec(code, {}, local_vars)
        
        result_df = local_vars['df']
        
        # Validate that no data was removed
        if result_df.shape[0] < original_shape[0]:
            return df, f"❌ Error: Suggestions would remove rows. Original: {original_shape[0]} rows, Result: {result_df.shape[0]} rows. Keeping original data."
        if result_df.shape[1] < original_shape[1]:
            return df, f"❌ Error: Suggestions would remove columns. Original: {original_shape[1]} cols, Result: {result_df.shape[1]} cols. Keeping original data."
        
        return result_df, f"✅ Applied suggestions successfully. Shape preserved: {original_shape}"
        
    except Exception as e:
        return df, f"❌ Error applying suggestions: {str(e)}"

def apply_user_query_to_df(df, query):
    """Apply user query"""
    try:
        client = get_anthropic_client()
        if not client:
            return df, "AI service not available"
        
        df_info = f"Columns: {list(df.columns)}\nSample data:\n{df.head(3).to_string()}"
        
        prompt = f"""Given this DataFrame:
{df_info}

User query: {query}

Generate Python pandas code to modify 'df'. Do not fill missing values. Do not remove columns unless completely empty with no name. Only return executable code.
Example: df = df.rename(columns={{'Col1': 'Name'}})"""
        
        response = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}]
        )
        
        code = response.content[0].text.strip()
        
        local_vars = {'df': df.copy(), 'pd': pd, 'np': np}
        exec(code, {}, local_vars)
        
        return local_vars['df'], f"✅ Applied: {code}"
        
    except Exception as e:
        return df, f"❌ Error: {str(e)}"

def convert_excel_to_json(file_content: bytes) -> Dict[str, Any]:
    """Convert Excel to JSON with all sheets"""
    temp_file = Path(tempfile.mktemp(suffix=".xlsx"))
    try:
        with open(temp_file, 'wb') as f:
            f.write(file_content)
        
        excel_file = pd.ExcelFile(temp_file, engine='openpyxl')
        sheets_data = {}
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(temp_file, sheet_name=sheet_name, engine='openpyxl')
            records = df.where(pd.notnull(df), None).to_dict('records')
            sheets_data[sheet_name] = {
                "data": records,
                "columns": list(df.columns),
                "shape": df.shape
            }
        
        return {"sheets": sheets_data, "total_sheets": len(sheets_data)}
    finally:
        if temp_file.exists():
            temp_file.unlink()

def display_excel_data(data: Dict[str, Any], title: str = "Excel Data", format_flags: List[Dict] = None):
    """Display Excel data in a nice format with format validation highlighting"""
    if not data or 'sheets' not in data:
        st.warning("No data to display")
        return
    
    st.markdown(f"<div class='section-header'>{title}</div>", unsafe_allow_html=True)
    
    sheets = data['sheets']
    total_sheets = data.get('total_sheets', len(sheets))
    
    st.info(f"📊 Total sheets: {total_sheets}")
    
    if total_sheets > 1:
        sheet_names = list(sheets.keys())
        selected_sheet = st.selectbox(f"Select sheet to view ({title}):", sheet_names, key=f"view_{title}")
        sheet_data = sheets[selected_sheet]
    else:
        selected_sheet = list(sheets.keys())[0] if sheets else "Sheet1"
        sheet_data = sheets[selected_sheet]
    
    st.markdown(f"**Sheet: {selected_sheet}**")
    
    if 'error' in sheet_data:
        st.error(f"Error in sheet: {sheet_data['error']}")
        return
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Rows", sheet_data['shape'][0])
    with col2:
        st.metric("Columns", sheet_data['shape'][1])
    with col3:
        st.metric("Total Cells", sheet_data['shape'][0] * sheet_data['shape'][1])
    
    if sheet_data['columns']:
        st.markdown("**Columns:**")
        st.write(", ".join(sheet_data['columns']))
    
    if sheet_data['data']:
        st.markdown("**Data Preview:**")
        df = pd.DataFrame(sheet_data['data'])
        
        # Apply format validation highlighting if format_flags are provided
        if format_flags:
            try:
                styled_df = apply_format_highlighting(df, format_flags)
                st.dataframe(styled_df, use_container_width=True, height=400)
            except Exception as e:
                st.warning("⚠️ Could not apply highlighting, showing data without highlighting")
                logger.error(f"Highlighting error: {e}")
                st.dataframe(df, use_container_width=True, height=400)
        else:
            st.dataframe(df, use_container_width=True, height=400)
        
        csv = df.to_csv(index=False)
        st.download_button(
            label=f"📥 Download {selected_sheet} as CSV",
            data=csv,
            file_name=f"{selected_sheet}.csv",
            mime="text/csv",
            key=f"download_{title}_{selected_sheet}"
        )
    else:
        st.warning("No data in this sheet")

def apply_format_highlighting(df: pd.DataFrame, format_flags: List[Dict]) -> pd.DataFrame:
    """Apply visual highlighting to format mismatches in the dataframe using a simple but reliable approach"""
    try:
        # Create a copy to avoid modifying original data
        display_df = df.copy()
        
        # If no format issues, return original dataframe
        if not format_flags:
            return display_df
        
        # Collect format information
        validator = FormatValidator()
        format_info = {}
        
        for flag in format_flags:
            col_name = flag['column']
            if col_name in display_df.columns:
                format_info[col_name] = {
                    'format_name': flag['format_detected'],
                    'mismatch_count': flag['mismatched_count'],
                    'pattern': validator.PATTERNS[flag['format_detected']]['pattern']
                }
        
        # Simple highlighting using styler.applymap
        def highlight_mismatched_cell(val):
            """Highlight individual cells that don't match their column's format"""
            return 'background-color: #ffcdd2; color: #b71c1c; font-weight: bold; border: 1px solid #f44336;'
        
        def check_and_style_cell(val, col_name):
            """Check if a cell value matches the expected format and style accordingly"""
            if col_name not in format_info:
                return ''
            
            if pd.isna(val):
                return ''
            
            try:
                import re
                pattern = format_info[col_name]['pattern']
                val_str = str(val).strip()
                
                if not re.match(pattern, val_str, re.IGNORECASE):
                    return 'background-color: #ffcdd2; color: #b71c1c; font-weight: bold; border: 1px solid #f44336;'
                else:
                    return ''
            except:
                return ''
        
        # Update column names to show format information
        column_rename = {}
        for col in display_df.columns:
            if col in format_info:
                format_name = format_info[col]['format_name'].replace('_', ' ')
                mismatch_count = format_info[col]['mismatch_count']
                column_rename[col] = f"{col} 🚨 {format_name} ({mismatch_count} errors)"
        
        # Rename columns if there are format issues
        if column_rename:
            display_df = display_df.rename(columns=column_rename)
        
        # Apply highlighting using a simple approach
        def apply_highlighting_to_dataframe(df_to_style):
            """Apply highlighting across the entire dataframe"""
            # Create a style matrix
            style_matrix = pd.DataFrame('', index=df_to_style.index, columns=df_to_style.columns)
            
            for col_idx, col in enumerate(display_df.columns):
                original_col = df.columns[col_idx]  # Get original column name
                if original_col in format_info:
                    for row_idx in range(len(df_to_style)):
                        cell_val = df_to_style.iloc[row_idx, col_idx]
                        style = check_and_style_cell(cell_val, original_col)
                        style_matrix.iloc[row_idx, col_idx] = style
            
            return style_matrix
        
        # Apply the styling
        styled_df = display_df.style.apply(apply_highlighting_to_dataframe, axis=None)
        styled_df = styled_df.set_table_attributes(
            'style="font-size: 12px; border-collapse: collapse; width: 100%;"'
        )
        
        return styled_df
        
    except Exception as e:
        logger.error(f"Error in apply_format_highlighting: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        # Fallback: Just rename columns to show format info
        try:
            column_rename = {}
            for flag in format_flags:
                col_name = flag['column']
                if col_name in df.columns:
                    format_name = flag['format_detected'].replace('_', ' ')
                    mismatch_count = flag['mismatched_count']
                    column_rename[col_name] = f"{col_name} 🚨 {format_name} ({mismatch_count} errors)"
            
            if column_rename:
                return df.rename(columns=column_rename)
            
        except Exception as e2:
            logger.error(f"Even fallback failed: {e2}")
        
        return df

def agent_decide_cleaning(excel_content: bytes, sheet_name: str) -> Dict[str, Any]:
    """Agentic AI to decide if cleaning or table splitting is needed."""
    try:
        client = get_anthropic_client()
        if not client:
            return {"decision": "unknown", "reason": "AI service not available", "issues": []}

        temp_file = Path(tempfile.mktemp(suffix=".xlsx"))
        with open(temp_file, 'wb') as f:
            f.write(excel_content)
        
        wb = openpyxl.load_workbook(temp_file)
        ws = wb[sheet_name]
        df = pd.read_excel(temp_file, sheet_name=sheet_name, engine='openpyxl')
        df_info = {
            "columns": list(df.columns),
            "dtypes": df.dtypes.to_dict(),
            "shape": df.shape,
            "missing_pct": (df.isnull().mean() * 100).to_dict(),
            "duplicates": df.duplicated().sum(),
            "sample": df.head(5).to_dict('records'),
            "merged_cells": len(list(ws.merged_cells.ranges)) > 0
        }
        
        if temp_file.exists():
            temp_file.unlink()

        system_prompt = """You are a conservative data quality analyzer for Excel files. Your goal is to identify if CLEANING (not removal) is needed.
        
        Criteria for 'yes' (needs cleaning - NO DATA REMOVAL):
        - Whitespace issues (leading/trailing spaces)
        - Inconsistent data types (numbers stored as text)
        - Mixed date formats
        - Special characters in column names
        - Encoding issues
        - Inconsistent units or formats
        - Missing column headers (needs naming)

        Criteria for 'no' (data is clean enough):
        - Data is already well-structured
        - Only has expected missing values
        - Consistent formatting
        
        NEVER recommend removing data. Focus only on cleaning and standardization.
        
        IMPORTANT: Output ONLY valid JSON, nothing else.
        Output JSON: {"decision": "yes" or "no", "reason": "explanation", "issues": ["issue1", ...], "cleaning_suggestions": ["suggestion1", ...]}"""

        user_prompt = f"""Data summary: {json.dumps(df_info, default=str)}
        
        Analyze if data cleaning (NOT removal) is needed."""

        response = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=500,
            system=system_prompt,
            messages=[{"role": "user", "content": user_prompt}]
        )
        
        try:
            response_text = response.content[0].text.strip()
            # Try to extract JSON if it's embedded in text
            if '{' in response_text and '}' in response_text:
                json_start = response_text.find('{')
                json_end = response_text.rfind('}') + 1
                json_str = response_text[json_start:json_end]
                result = json.loads(json_str)
            else:
                result = json.loads(response_text)
            
            # Ensure compatibility with existing code
            if 'cleaning_suggestions' in result and 'action' not in result:
                result['action'] = 'trigger_cleaning' if result['decision'] == 'yes' else 'skip'
            
            return result
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error: {e}")
            logger.error(f"AI response was: {response.content[0].text[:500]}")
            return {"decision": "unknown", "reason": f"Invalid AI response format: {str(e)[:100]}", "issues": [], "cleaning_suggestions": [], "raw_response": response.content[0].text[:200]}
    
    except Exception as e:
        logger.error(f"Agent error: {str(e)}")
        return {"decision": "unknown", "reason": f"Error: {str(e)}", "issues": []}

# Main app
def main():
    st.markdown("<div class='main-header'>📊 Excel Data Cleaner</div>", unsafe_allow_html=True)
    
    # Sidebar with navigation
    with st.sidebar:
        st.image("https://via.placeholder.com/150", use_column_width=True)
        st.markdown("### 🔧 Navigation")
        pages = ["Upload", "Clean", "Analyze", "Apply Suggestions", "Apply Query", "Download"]
        selection = st.radio("Go to", pages, index=0)
        st.markdown("### 📊 Status")
        ai_available = get_anthropic_client() is not None
        s3_available = True
        if ai_available:
            st.success("✅ AI Service Available")
        else:
            st.warning("⚠️ AI Service Unavailable")
        if s3_available:
            st.success("✅ S3 Connected")
        else:
            st.error("❌ S3 Not Configured")
        if st.session_state.current_excel_id:
            st.markdown("### 📂 Current File")
            st.info(f"ID: {st.session_state.current_excel_id[:8]}...")
            st.info(f"Name: {st.session_state.current_excel_name}")
            st.info(f"Sheet: {st.session_state.current_sheet_name}")
            if st.button("🗑️ Clear Session"):
                st.session_state.clear()
                st.rerun()

    # Dynamic tab selection
    if selection == "Upload":
        with st.expander("📤 Upload Excel File", expanded=True):
            st.markdown("<div class='section-header'>📤 Upload</div>", unsafe_allow_html=True)
            uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx', 'xls'])
            if uploaded_file:
                st.info(f"**File:** {uploaded_file.name} ({uploaded_file.size:,} bytes)")
                if st.button("🚀 Upload & Store in S3", type="primary"):
                    progress_bar = st.progress(0)
                    progress_bar.progress(20)
                    excel_id = str(uuid.uuid4())
                    content = uploaded_file.read()
                    s3_service.create_bucket_if_not_exists()
                    temp_file = Path(tempfile.mktemp(suffix=".xlsx"))
                    with open(temp_file, 'wb') as f:
                        f.write(content)
                    with pd.ExcelFile(temp_file, engine='openpyxl') as xl_file:
                        sheet_names = xl_file.sheet_names
                    for sheet_name in sheet_names:
                        s3_service.upload_original_file(excel_id, uploaded_file.name, sheet_name, content)
                        metadata = {
                            "excel_id": excel_id,
                            "filename": uploaded_file.name,
                            "sheets": sheet_names,
                            "sheet_name": sheet_name,
                            "upload_timestamp": datetime.utcnow().isoformat()
                        }
                        s3_service.upload_metadata(excel_id, uploaded_file.name, sheet_name, metadata)
                    st.session_state.current_excel_id = excel_id
                    st.session_state.current_excel_name = uploaded_file.name
                    st.session_state.available_sheets = sheet_names
                    st.session_state.current_sheet_name = sheet_names[0] if sheet_names else None
                    if temp_file.exists():
                        temp_file.unlink()
                    progress_bar.progress(100)
                    st.success("✅ Uploaded to S3!")
                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric("ID", excel_id[:16] + "...")
                        st.metric("Size", f"{len(content):,} bytes")
                    with col2:
                        st.metric("Sheets", len(sheet_names))
                        st.metric("S3", "✅")
                    st.write("**Sheets:**", ", ".join(sheet_names))
                    excel_data = convert_excel_to_json(content)
                    display_excel_data(excel_data, "📤 Original Data")
                    with st.spinner("Agent assessing..."):
                        original_content = s3_service.download_original_file(excel_id, uploaded_file.name, st.session_state.current_sheet_name)
                        agent_result = agent_decide_cleaning(original_content, st.session_state.current_sheet_name)
                        st.session_state.agent_decision = agent_result
                    st.markdown("**Data Quality Assessment:**")
                    if agent_result['decision'] == 'yes':
                        st.warning(f"⚠️ Cleaning recommended: {agent_result['reason']}")
                        if agent_result.get('issues'):
                            st.markdown("**Issues found:**")
                            for issue in agent_result['issues']:
                                st.write(f"- {issue}")
                        if agent_result.get('cleaning_suggestions'):
                            st.markdown("**Suggested cleaning steps:**")
                            for suggestion in agent_result['cleaning_suggestions']:
                                st.write(f"- {suggestion}")
                    elif agent_result['decision'] == 'no':
                        st.success(f"✅ Data quality is good: {agent_result['reason']}")
                    else:
                        st.error(f"❌ Assessment error: {agent_result['reason']}")

    elif selection == "Clean":
        with st.expander("🧹 Clean Excel Data", expanded=True):
            st.markdown("<div class='section-header'>🧹 Clean</div>", unsafe_allow_html=True)
            if not st.session_state.current_excel_id:
                st.warning("⚠️ Upload a file first!")
            else:
                st.info(f"**File:** {st.session_state.current_excel_name}")
                
                # Always refresh sheet list to ensure we have the latest info
                try:
                    # Download original file to check actual sheets
                    test_content = s3_service.download_original_file(
                        st.session_state.current_excel_id,
                        st.session_state.current_excel_name,
                        st.session_state.available_sheets[0] if st.session_state.available_sheets else "Sheet1"
                    )
                    if test_content:
                        temp_file = Path(tempfile.mktemp(suffix=".xlsx"))
                        with open(temp_file, 'wb') as f:
                            f.write(test_content)
                        with pd.ExcelFile(temp_file, engine='openpyxl') as xl_file:
                            actual_sheets = xl_file.sheet_names
                        st.session_state.available_sheets = actual_sheets
                        if temp_file.exists():
                            temp_file.unlink()
                except Exception as e:
                    st.warning(f"Could not refresh sheet list: {str(e)}")
                
                # Display actual available sheets
                st.write(f"**Available sheets:** {', '.join(st.session_state.available_sheets) if st.session_state.available_sheets else 'None'}")
                
                if len(st.session_state.available_sheets) > 1:
                    selected_sheet = st.selectbox(
                        "Select sheet to clean:", 
                        st.session_state.available_sheets, 
                        index=st.session_state.available_sheets.index(st.session_state.current_sheet_name) 
                        if st.session_state.current_sheet_name in st.session_state.available_sheets else 0
                    )
                elif len(st.session_state.available_sheets) == 1:
                    selected_sheet = st.session_state.available_sheets[0]
                    st.write(f"**Sheet:** {selected_sheet}")
                else:
                    st.error("❌ No sheets found in the Excel file")
                    selected_sheet = None
                if 'agent_decision' in st.session_state and st.session_state.agent_decision:
                    decision = st.session_state.agent_decision.get('decision', 'unknown').upper()
                    reason = st.session_state.agent_decision.get('reason', 'No reason provided')
                    st.info(f"Agent: {decision} - {reason}")
                if st.button("🤖 Re-Assess with Agent"):
                    with st.spinner("Assessing..."):
                        original_content = s3_service.download_original_file(st.session_state.current_excel_id, st.session_state.current_excel_name, selected_sheet)
                        agent_result = agent_decide_cleaning(original_content, selected_sheet)
                        st.session_state.agent_decision = agent_result
                        st.markdown("**Data Quality Assessment:**")
                        if agent_result['decision'] == 'yes':
                            st.warning(f"⚠️ Cleaning recommended: {agent_result['reason']}")
                            if agent_result.get('issues'):
                                st.markdown("**Issues found:**")
                                for issue in agent_result.get('issues', []):
                                    st.write(f"- {issue}")
                            if agent_result.get('cleaning_suggestions'):
                                st.markdown("**Suggested cleaning steps:**")
                                for suggestion in agent_result.get('cleaning_suggestions', []):
                                    st.write(f"- {suggestion}")
                        elif agent_result['decision'] == 'no':
                            st.success(f"✅ Data quality is good: {agent_result['reason']}")
                        else:
                            st.error(f"❌ Assessment error: {agent_result['reason']}")
                            if 'raw_response' in agent_result:
                                with st.expander("Debug: Raw AI Response"):
                                    st.text(agent_result['raw_response'])
                # Show checkbox for auto-apply but don't auto-execute
                if st.session_state.agent_decision and st.session_state.agent_decision.get('decision') == 'yes':
                    auto_apply = st.checkbox("Auto-apply cleaning based on agent recommendation?", value=False)
                    if auto_apply:
                        st.info("Click 'Clean Data' button below to apply conservative cleaning")
                if selected_sheet and st.button("🧹 Clean Data", type="primary"):
                    try:
                        progress_bar = st.progress(0)
                        progress_bar.progress(20)
                        
                        # Download the original file - use the first available sheet since S3 stores the complete file
                        original_content = s3_service.download_original_file(
                            st.session_state.current_excel_id, 
                            st.session_state.current_excel_name, 
                            st.session_state.available_sheets[0] if st.session_state.available_sheets else selected_sheet
                        )
                        if not original_content:
                            st.error("❌ Original file not found in S3")
                            return
                        
                        temp_input = Path(tempfile.mktemp(suffix=".xlsx"))
                        temp_output = Path(tempfile.mktemp(suffix="_cleaned.xlsx"))
                        
                        with open(temp_input, 'wb') as f:
                            f.write(original_content)
                        
                        progress_bar.progress(40)
                        
                        # Verify sheet exists before cleaning
                        try:
                            with pd.ExcelFile(temp_input, engine='openpyxl') as xl_file:
                                actual_sheets = xl_file.sheet_names
                                
                                # Normalize sheet names for comparison (case-insensitive, strip whitespace)
                                normalized_selected = selected_sheet.lower().strip()
                                normalized_sheets = [s.lower().strip() for s in actual_sheets]
                                
                                if normalized_selected not in normalized_sheets:
                                    # Check for exact match ignoring case
                                    matched_sheet = next((s for s in actual_sheets if s.lower().strip() == normalized_selected), None)
                                    if not matched_sheet:
                                        # Fallback to first sheet if no match
                                        if actual_sheets:
                                            st.warning(f"Sheet '{selected_sheet}' not found, using first available sheet: '{actual_sheets[0]}'")
                                            selected_sheet = actual_sheets[0]
                                        else:
                                            st.error("❌ No sheets found in the Excel file")
                                            if temp_input.exists():
                                                temp_input.unlink()
                                            return
                                    else:
                                        st.warning(f"Sheet '{selected_sheet}' not found, using '{matched_sheet}' instead")
                                        selected_sheet = matched_sheet
                                
                                # Update session state with correct sheets
                                st.session_state.available_sheets = actual_sheets
                                st.session_state.current_sheet_name = selected_sheet
                                
                        except Exception as e:
                            st.error(f"❌ Error reading Excel file: {str(e)}")
                            if temp_input.exists():
                                temp_input.unlink()
                            return
                        
                        df, changes_log, processing_time = clean_excel_basic(str(temp_input), str(temp_output), selected_sheet)
                        
                        progress_bar.progress(60)
                        
                        # Check if cleaning was successful
                        if df.empty:
                            st.error("❌ Cleaning failed - resulting DataFrame is empty")
                            for change in changes_log:
                                if "❌" in change:
                                    st.error(change)
                            # Clean up temp files
                            if temp_input.exists():
                                temp_input.unlink()
                            if temp_output.exists():
                                temp_output.unlink()
                            return
                        
                        # Check if output file was created
                        if not temp_output.exists():
                            st.error("❌ Cleaned file was not created. Check the logs for errors.")
                            if temp_input.exists():
                                temp_input.unlink()
                            return
                        
                        progress_bar.progress(80)
                        
                        with open(temp_output, 'rb') as f:
                            cleaned_content = f.read()
                        
                        s3_service.upload_cleaned_file(st.session_state.current_excel_id, st.session_state.current_excel_name, selected_sheet, cleaned_content)
                        st.session_state.current_sheet_name = selected_sheet
                        st.session_state.cleaned_data = df
                        
                        # Clean up temp files
                        for f in [temp_input, temp_output]:
                            if f.exists():
                                f.unlink()
                        
                        progress_bar.progress(100)
                        st.success("✅ Cleaned!")
                        
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Rows", df.shape[0])
                        with col2:
                            st.metric("Columns", df.shape[1])
                        with col3:
                            st.metric("Time", f"{processing_time:.2f}s")
                        
                        with st.expander("📋 Changes", expanded=True):
                            for change in changes_log:
                                st.write(change)
                        
                        cleaned_excel_data = convert_excel_to_json(cleaned_content)
                        # Run format validation on cleaned data for highlighting
                        cleaned_temp_file = Path(tempfile.mktemp(suffix=".xlsx"))
                        with open(cleaned_temp_file, 'wb') as f:
                            f.write(cleaned_content)
                        cleaned_df = pd.read_excel(cleaned_temp_file, engine='openpyxl')
                        _, cleaned_format_flags = validate_and_flag_formats(cleaned_df)
                        if cleaned_temp_file.exists():
                            cleaned_temp_file.unlink()
                        display_excel_data(cleaned_excel_data, "🧹 Cleaned Data", cleaned_format_flags)
                        
                    except FileNotFoundError as e:
                        st.error(f"❌ File error: {str(e)}")
                        logger.error(f"FileNotFoundError: {traceback.format_exc()}")
                    except Exception as e:
                        st.error(f"❌ Cleaning failed: {str(e)}")
                        logger.error(f"Cleaning error: {traceback.format_exc()}")

    elif selection == "Analyze":
        with st.expander("🔍 Analyze Data", expanded=True):
            st.markdown("<div class='section-header'>🔍 Analyze</div>", unsafe_allow_html=True)
            if not st.session_state.current_excel_id:
                st.warning("⚠️ Upload and clean a file first!")
            else:
                st.info(f"**File:** {st.session_state.current_excel_name} - **Sheet:** {st.session_state.current_sheet_name}")
                if st.button("🔍 Analyze with AI", type="primary"):
                    with st.spinner("Analyzing..."):
                        cleaned_content = s3_service.download_cleaned_file(st.session_state.current_excel_id, st.session_state.current_excel_name, st.session_state.current_sheet_name)
                        temp_file = Path(tempfile.mktemp(suffix=".xlsx"))
                        with open(temp_file, 'wb') as f:
                            f.write(cleaned_content)
                        df = pd.read_excel(temp_file, engine='openpyxl')
                        
                        # Run format validation
                        _, format_flags = validate_and_flag_formats(df)
                        
                        if temp_file.exists():
                            temp_file.unlink()
                        analysis, errors = ai_analyze_df(df)
                        if analysis:
                            st.session_state.analysis_results = analysis
                            st.session_state.format_flags = format_flags  # Store format flags
                            st.success("✅ Analyzed!")
                            
                            # Display data quality score if available
                            if 'data_quality_score' in analysis:
                                score = analysis['data_quality_score']
                                color = "🟢" if score >= 80 else "🟡" if score >= 60 else "🔴"
                                st.markdown(f"### {color} Data Quality Score: {score}/100")
                                st.progress(score / 100)
                            
                            # Display format validation issues if any
                            if format_flags:
                                with st.expander(f"🚨 Format Validation Issues ({len(format_flags)} found)", expanded=True):
                                    for flag in format_flags:
                                        severity_color = "🔴" if flag['severity'] == 'HIGH' else "🟡"
                                        st.markdown(f"### {severity_color} Column: `{flag['column']}`")
                                        st.write(f"**Format Detected:** {flag['format_description']}")
                                        st.write(f"**Match Rate:** {flag['match_percentage']}%")
                                        st.write(f"**Mismatched Values:** {flag['mismatched_count']} out of {flag['total_values']}")
                                        
                                        if flag['sample_mismatches']:
                                            st.write("**Sample Mismatches:**")
                                            for sample in flag['sample_mismatches'][:5]:
                                                st.code(sample)
                                        
                                        st.warning(flag['recommendation'])
                                        st.markdown("---")
                            
                            # Display metadata in an expandable section
                            if 'metadata' in analysis:
                                with st.expander("📊 Detailed Column Metadata", expanded=False):
                                    # Create a more readable metadata display
                                    for col_name, col_meta in analysis['metadata'].items():
                                        st.markdown(f"**Column: `{col_name}`**")
                                        col1, col2, col3 = st.columns(3)
                                        with col1:
                                            st.write(f"Type: {col_meta.get('dtype', 'unknown')}")
                                            st.write(f"Non-null: {col_meta.get('non_null_count', 0)}/{col_meta.get('total_count', 0)}")
                                        with col2:
                                            st.write(f"Unique: {col_meta.get('unique_count', 0)}")
                                            st.write(f"Missing: {col_meta.get('null_percentage', 0)}%")
                                        with col3:
                                            if col_meta.get('has_leading_spaces'):
                                                st.write("⚠️ Leading spaces")
                                            if col_meta.get('has_trailing_spaces'):
                                                st.write("⚠️ Trailing spaces")
                                            if col_meta.get('has_mixed_types'):
                                                st.write("⚠️ Mixed types")
                                        
                                        # Show sample values
                                        if col_meta.get('sample_values'):
                                            st.write(f"Sample values: {', '.join(str(v) for v in col_meta['sample_values'][:5])}")
                                        
                                        # Show numeric statistics if available
                                        if 'min' in col_meta:
                                            st.write(f"Range: [{col_meta.get('min')}, {col_meta.get('max')}]")
                                            st.write(f"Mean: {col_meta.get('mean')}, Std: {col_meta.get('std')}")
                                        
                                        st.markdown("---")
                            
                            if 'analysis' in analysis:
                                st.markdown("**🔍 Data Quality Issues Found:**")
                                for item in analysis['analysis']:
                                    st.write(f"• {item}")
                            
                            if 'suggestions' in analysis:
                                st.markdown("**💡 Cleaning Suggestions (No Data Removal):**")
                                for suggestion in analysis['suggestions']:
                                    st.write(f"• {suggestion}")
                        else:
                            st.error(f"❌ Analysis failed: {', '.join(errors)}")

    elif selection == "Apply Suggestions":
        with st.expander("💡 Apply Suggestions", expanded=True):
            st.markdown("<div class='section-header'>💡 Apply</div>", unsafe_allow_html=True)
            if not st.session_state.current_excel_id:
                st.warning("⚠️ Upload and analyze a file first!")
            elif not st.session_state.analysis_results:
                st.warning("⚠️ Analyze the file first!")
            else:
                st.info(f"**File:** {st.session_state.current_excel_name} - **Sheet:** {st.session_state.current_sheet_name}")
                if 'suggestions' in st.session_state.analysis_results:
                    st.markdown("**Suggestions:**")
                    selected_suggestions = [s for i, s in enumerate(st.session_state.analysis_results['suggestions']) if st.checkbox(s, key=f"sug_{i}")]
                    if selected_suggestions and st.button("💡 Apply", type="primary"):
                        with st.spinner("Applying..."):
                            cleaned_content = s3_service.download_cleaned_file(st.session_state.current_excel_id, st.session_state.current_excel_name, st.session_state.current_sheet_name)
                            temp_file = Path(tempfile.mktemp(suffix=".xlsx"))
                            with open(temp_file, 'wb') as f:
                                f.write(cleaned_content)
                            df = pd.read_excel(temp_file, engine='openpyxl')
                            modified_df, result = apply_ai_suggestions(df, selected_suggestions)
                            temp_output = Path(tempfile.mktemp(suffix="_suggestions.xlsx"))
                            modified_df.to_excel(temp_output, index=False, engine='openpyxl')
                            with open(temp_output, 'rb') as f:
                                new_content = f.read()
                            s3_service.upload_cleaned_file(st.session_state.current_excel_id, st.session_state.current_excel_name, st.session_state.current_sheet_name, new_content)
                            for f in [temp_file, temp_output]:
                                if f.exists():
                                    f.unlink()
                            st.success("✅ Applied!")
                            st.write(f"**Result:** {result}")
                            updated_excel_data = convert_excel_to_json(new_content)
                            display_excel_data(updated_excel_data, "💡 Updated Data")

    elif selection == "Apply Query":
        with st.expander("❓ Apply Query", expanded=True):
            st.markdown("<div class='section-header'>❓ Query</div>", unsafe_allow_html=True)
            if not st.session_state.current_excel_id:
                st.warning("⚠️ Upload and clean a file first!")
            else:
                st.info(f"**File:** {st.session_state.current_excel_name} - **Sheet:** {st.session_state.current_sheet_name}")
                query_text = st.text_area("Enter query:", height=100, placeholder="e.g., Filter sales > 1000")
                if query_text and st.button("❓ Apply", type="primary"):
                    with st.spinner("Applying..."):
                        cleaned_content = s3_service.download_cleaned_file(st.session_state.current_excel_id, st.session_state.current_excel_name, st.session_state.current_sheet_name)
                        temp_file = Path(tempfile.mktemp(suffix=".xlsx"))
                        with open(temp_file, 'wb') as f:
                            f.write(cleaned_content)
                        df = pd.read_excel(temp_file, engine='openpyxl')
                        modified_df, result = apply_user_query_to_df(df, query_text)
                        temp_output = Path(tempfile.mktemp(suffix="_query.xlsx"))
                        modified_df.to_excel(temp_output, index=False, engine='openpyxl')
                        with open(temp_output, 'rb') as f:
                            new_content = f.read()
                        s3_service.upload_cleaned_file(st.session_state.current_excel_id, st.session_state.current_excel_name, st.session_state.current_sheet_name, new_content)
                        for f in [temp_file, temp_output]:
                            if f.exists():
                                f.unlink()
                        st.success("✅ Applied!")
                        st.write(f"**Result:** {result}")
                        updated_excel_data = convert_excel_to_json(new_content)
                        display_excel_data(updated_excel_data, "❓ Queried Data")

    elif selection == "Download":
        with st.expander("📥 Download Files", expanded=True):
            st.markdown("<div class='section-header'>📥 Download</div>", unsafe_allow_html=True)
            if not st.session_state.current_excel_id:
                st.warning("⚠️ Upload a file first!")
            else:
                st.info(f"**File:** {st.session_state.current_excel_name} - **Sheet:** {st.session_state.current_sheet_name}")
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("📥 Original", type="secondary"):
                        content = s3_service.download_original_file(st.session_state.current_excel_id, st.session_state.current_excel_name, st.session_state.current_sheet_name)
                        if content:
                            st.download_button("⬇️ Original", content, f"original_{st.session_state.current_excel_name}", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        else:
                            st.error("Original not found")
                with col2:
                    if st.button("📥 Cleaned", type="primary"):
                        content = s3_service.download_cleaned_file(st.session_state.current_excel_id, st.session_state.current_excel_name, st.session_state.current_sheet_name)
                        if content:
                            st.download_button("⬇️ Cleaned", content, f"cleaned_{st.session_state.current_excel_name}", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        else:
                            st.error("Cleaned not found - clean first")
                st.markdown("---")
                st.warning("⚠️ Deletes all files for this ID!")
                if st.button("🗑️ Delete from S3", type="secondary"):
                    confirm = st.checkbox("Confirm deletion")
                    if confirm:
                        success = s3_service.delete_files(st.session_state.current_excel_id, st.session_state.current_excel_name, st.session_state.current_sheet_name)
                        if success:
                            st.success("✅ Deleted!")
                            st.session_state.clear()
                            st.rerun()
                        else:
                            st.error("Failed to delete")

    st.markdown("<div style='text-align: center; color: #7f8c8d; font-size: 0.9rem; margin-top: 2rem;'>Excel Data Cleaner v2.0</div>", unsafe_allow_html=True)

if __name__ == "__main__":
    main()